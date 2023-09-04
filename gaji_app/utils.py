from openpyxl import load_workbook
from .models import Absen, User, Karyawan, Divisi
import calendar
from django.db.models import Count
from datetime import datetime
from workdays import workday

def absence_excel_file(file):
    workbook = load_workbook(file)
    worksheet = workbook.active

    success_count = 0
    error_messages = []

    for row in worksheet.iter_rows(min_row=2):
        try:
            # Assuming columns are in A, B, C order in the Excel file
            column_a_value = row[0].value
            column_b_value = row[1].value
            column_c_value = row[2].value

            # Create a record in the database
            karyawan = Karyawan.objects.get(nip=column_c_value)
            Absen.objects.create(
                    nip=karyawan,
                    tanggal=column_a_value,
                    keterangan=column_b_value,
                )
            success_count += 1
        except Exception as e:
            error_messages.append(str(e))

    return success_count, error_messages

def user_excel_file(file):
    workbook = load_workbook(file)
    worksheet = workbook.active

    success_count = 0
    error_messages = []

    for row in worksheet.iter_rows(min_row=2):
        try:
            # Assuming columns are in A, B, C order in the Excel file
            column_a_value = row[0].value
            column_b_value = row[1].value
            column_c_value = row[2].value
            nip = row[3].value
            address = row[4].value
            phone_number = row[5].value
            position = row[6].value
            jenis_karyawan = row[7].value
            divisi_id = row[8].value
            

            # Create a record in the database
            user = User.objects.create(nama=column_a_value, email=column_b_value)
            user.set_password(column_c_value)
            user.save()
            User.objects.get(nama=column_a_value)
            Karyawan.objects.create(user=user, nama=column_a_value, nip=nip, address=address, phone_number=phone_number, position=position, jenis_karyawan=jenis_karyawan, divisi_id=divisi_id)
            success_count += 1
        except Exception as e:
            error_messages.append(str(e))

    return success_count, error_messages

def divisi_excel_file(file):
    workbook = load_workbook(file)
    worksheet = workbook.active

    success_count = 0
    error_messages = []

    for row in worksheet.iter_rows(min_row=2):
        try:
            # Assuming columns are in A, B, C order in the Excel file
            column_a_value = row[0].value
            column_b_value = row[1].value

            # Create a record in the database
            Divisi.objects.create(nama_divisi=column_b_value)
            success_count += 1
        except Exception as e:
            error_messages.append(str(e))

    return success_count, error_messages

def riwayat_absen(nip):
        today = datetime.today()
        current_year = today.year
        current_month = 8
        last_day = calendar.monthrange(current_year, current_month)[1]
        
        start_date = today.replace(day=1)
        end_date = today.replace(day=last_day)
        
        riwayat_masuk = Absen.objects.filter(
            nip=nip,
            tanggal__range=(start_date, end_date)
        ).exclude(tanggal__week_day=calendar.SUNDAY)
        
        hitung_absen = riwayat_masuk.values('keterangan').annotate(count=Count('keterangan'))
        
        return hitung_absen
    
def hari_kerja(year, month):
    last_day = calendar.monthrange(year, month)[1]
    first_day = datetime(year, month, 1)
    last_day = datetime(year, month, last_day)
    
    total_hari_kerja = 0
    current_day = first_day
    while current_day <= last_day:
        if current_day.weekday() <6:
            total_hari_kerja += 1
        current_day = workday(current_day, 1)
        
    return total_hari_kerja

def basic_salary(posisi, divisi):
    if posisi == "staff" and divisi == 1:
        upah_bulan = 2400000
    elif posisi == "direksi" and divisi == 2:
        upah_bulan = 5500000
    elif posisi == "manajer" and divisi == 2:
        upah_bulan = 5000000
    elif posisi == "staff" and divisi == 3:
        upah_bulan = 3800000
    elif posisi == "manajer" and divisi == 3:
        upah_bulan = 4200000
    elif posisi == "staff" and divisi == 4:
        upah_bulan = 3800000
    elif posisi == "manajer" and divisi == 4:
        upah_bulan = 4200000
    elif posisi == "staff" and divisi == 5:
        upah_bulan = 3800000
    elif posisi == "manajer" and divisi == 5:
        upah_bulan = 4200000
    elif posisi == "staff" and divisi == 6:
        upah_bulan = 3800000
    elif posisi == "manajer" and divisi == 6:
        upah_bulan = 4200000
    elif posisi == "staff" and divisi == 7:
        upah_bulan = 3800000
    elif posisi == "manajer" and divisi == 7:
        upah_bulan = 4200000
    else: 
        upah_bulan =6000000
    
    return upah_bulan

def find_gaji_pokok(upah_bulan, total_hari_kerja, jml_hari_masuk):
    gaji = (upah_bulan / total_hari_kerja) * jml_hari_masuk
    gaji_pokok = gaji - (upah_bulan* 0.05)
    return gaji_pokok
    
def find_total_gaji(gaji_pokok, tunjangan_cuti, tunjangan_makan, tunjangan_transport):
    gaji_bersih_tahun = gaji_pokok * 12
    pph = (gaji_bersih_tahun - 48000000) * 0.05 / 12
    total_gaji = gaji_pokok - pph + tunjangan_cuti + tunjangan_makan + tunjangan_transport
    return total_gaji
